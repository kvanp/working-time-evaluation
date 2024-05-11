"""Load worktimes from a Excel sheet"""

from openpyxl import load_workbook
import worktime

class table:
    """List of worktime entrys from .xlsx file"""
    columns = {
        "date" : 1,
        "start" : 2,
        "break start" : 3,
        "break end" : 5,
        "end" : 6
    }

    def load(self, filename):
        """Load a Table from a Excel sheet

        Arguments:
        file_ -- path/filename to the logfile

        Return: List
        """
        wb = load_workbook(filename=filename)
        sheet = wb[wb.sheetnames[0]]
        offset = 5
        content = []
        date = None

        while offset < (sheet.max_row - 2):
            line = {}

            for k,v in self.columns.items():
                val = sheet.cell(offset, v).value

                # At 0 o'clock we get the wrong data type
                if not k == "date":
                    if isinstance(val, worktime.datetime.datetime):
                        val = val.time()
                    elif type(val) == str:
                        val = worktime.datetime.datetime.strptime(val, "%H:%M").time()
                else:
                    if val:
                        date = val

                if val:
                    line[k] = val

            if date and not "date" in line:
                line["date"] = date

            if line and "start" in line and "end" in line:
                content.append(line)

            offset += 1

        return content

class list(worktime.raw_list):
    def create(self, filename):
        """Create a new raw list

        Arguments:
        file_ -- path/filename to the logfile
        """
        self.list = table().load(filename)

    def append(self, filename):
        """Appand entrys to the raw list

        Arguments:
        file_ -- path/filename to the logfile
        """
        self.list += table().load(filename)

    def convert(self):
        """Convert the raw list to the worktime list type

        Return: worktime.list()
        """
        working_time_list = worktime.list()

        for entry in self.list:
            date = entry["date"]
            if "start" in entry:
                working_time_list.append(worktime.datetime.datetime.combine(date.date(), entry["start"]),
                                         worktime.enum_stamp_type.start_of_work)
            if "break start" in entry:
                working_time_list.append(worktime.datetime.datetime.combine(date.date(), entry["break start"]),
                                         worktime.enum_stamp_type.start_of_work_break)
            if "break end" in entry:
                working_time_list.append(worktime.datetime.datetime.combine(date.date(), entry["break end"]),
                                         worktime.enum_stamp_type.end_of_work_break)
            if "end" in entry:
                working_time_list.append(worktime.datetime.datetime.combine(date.date(), entry["end"]),
                                         worktime.enum_stamp_type.end_of_work)

        return working_time_list

