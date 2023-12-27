"""Load worktimes from a Excel sheet"""

import worktime
from openpyxl import load_workbook

class table:
    """List of worktime entrys from .xlsx file"""
    columns = {
        "date" : 1,
        "start" : 2,
        "break start" : 3,
        "break end" : 5,
        "end" : 6
    }

    def load(self, filename):
        """Create a new raw list

        Arguments:
        file_ -- path/filename to the logfile
        """
        wb = load_workbook(filename=filename)
        sheet = wb[wb.sheetnames[0]]
        offset = 5
        content = []
        while sheet.cell(offset, 1).value:
            line = {}
            for k,v in self.columns.items():
                val = sheet.cell(offset, v).value
                if val:
                    line[k] = val
            if line:
                content.append(line)
            offset += 1
            if offset > sheet.max_row:
                break
        return content

class list(worktime.raw_list):
    def create(self, filename):
        """Create a new raw list

        Arguments:
        file_ -- path/filename to the logfile
        """
        self.list = table().load(filename)

    def append(self, filename):
        """Appand entrys to the raw list

        Arguments:
        file_ -- path/filename to the logfile
        """
        self.list += table().load(filename)

    def convert(self):
        """Convert the raw list to the worktime list type"""
        for l in self.list:
            print(l)
        return worktime.list()

